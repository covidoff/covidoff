from flask import Flask, render_template_string, request, redirect
from openpyxl import load_workbook
import openpyxl as pxl
import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from bokeh.models.widgets import Button
from bokeh.models import CustomJS
from streamlit_bokeh_events import streamlit_bokeh_events
from io import StringIO
from PIL import Image
from gspread_dataframe import get_as_dataframe, set_with_dataframe

import gspread
from oauth2client.service_account import ServiceAccountCredentials

scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_name(
    'covidoff-ecef33b9fe0b.json', scope)
st.write("# **CovidOff View Data Portal**")
st.write("This is the Portal for all the Data Resources. These resources are filled by general people. Verification at utmostlevel is not guaranteed. Select your required service from the below dropdown")
selected = st.selectbox(
    "Select Service", ("Choose from the DropDown", "View Plasma Requirements", "View Plasma Donors", "View Oxygen Requirements", "View Oxygen Suppliers/Availability", "Search for Bed Availability", "Search for Remdesivr Distributors", "Other Resources"))
if selected == "Choose from the DropDown":
    pass
if selected == "View Plasma Requirements":

    gc = gspread.authorize(credentials)
    sht2 = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/1kcj_u4j9269CcdQGXuLzdU-WzrtWdR5vdOEwN7-46JM/edit#gid=0')
    worksheet = sht2.sheet1
    df2 = get_as_dataframe(worksheet)
    df2 = df2[df2['phone'].notna()]
    df2.rename(columns={'timestamp': 'Timestamp', 'name': 'Name / Contact Person', 'phone': 'Contact Number', 'bloodGroup': 'Blood Group', 'age': 'Age',
                        'gender': 'Gender', 'Cities': 'City / Area / Location', 'Category': 'State / Territory', 'selectDistrict': 'District', 'email': 'Email Id'}, inplace=True)
    df2.fillna(0)
    df = df2.to_excel('./plasmareqoutput.xlsx', sheet_name='Sheet1')
    exp = pxl.load_workbook('plasmareqoutput.xlsx')
    sheet = exp['Sheet1']
    st.write("## **Showing Plasma Requirements**")
    page_name = ['Choose Option', 'View Full Data',
                 'View Filtered Data']
    page = st.radio('Choose your preferred type', page_name)
    if page == "Choose Option":
        st.info("👆 You have to choose any of the above two Options.")
    elif page == "View Full Data":
        df2 = df2.iloc[::-1]
        st.write(df2.iloc[:, 1:11])
    elif page == "View Filtered Data":
        st.write(
            "Search the Plasma Requirements by District or Blood Group or Combined")
        option_t = st.checkbox('Blood Group')
        option_u = st.checkbox('District')
        if(option_t == True and option_u == True):
            district = st.selectbox(
                "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                    "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                    ))
            if district == "Not Listed":
                district = st.text_input('Enter Your District Name')
            bg = st.selectbox("Enter Blood Group", ("Select Blood Group",
                                                    "A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"))
            list = []
            for i in range(sheet.max_row+1, 2, -1):
                if (str(sheet.cell(row=i, column=5).value).upper() == bg.upper()) and (str(sheet.cell(row=i, column=9).value).upper() == district.upper()):
                    list.append(i-2)
            st.write(df2.iloc[list, 1:11])
            list = []
        if(option_t == True and option_u == False):

            bg = st.selectbox("Enter Blood Group", ("Select Blood Group",
                                                    "A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"))
            list = []
            for i in range(sheet.max_row+1, 2, -1):
                if (str(sheet.cell(row=i, column=5).value).upper() == bg.upper()):
                    list.append(i-2)
            st.write(df2.iloc[list, 1:11])
            list = []
        if(option_t == False and option_u == True):
            district = st.selectbox(
                "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                    "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                    ))
            if district == "Not Listed":
                district = st.text_input('Enter Your District Name')
            list = []
            for i in range(sheet.max_row+1, 2, -1):
                if (str(sheet.cell(row=i, column=9).value).upper() == district.upper()):
                    list.append(i-2)
            st.write(df2.iloc[list, 1:11])
            list = []

if selected == "View Plasma Donors":
    gc = gspread.authorize(credentials)
    sht2 = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/1RR3Iu0fyzzdGkhUsOReijAIRaL1q0f6a20sHeqvQgmY/edit#gid=0')
    worksheet = sht2.sheet1
    df2 = get_as_dataframe(worksheet)
    df2 = df2[df2['phone'].notna()]
    df2.rename(columns={'timestamp': 'Timestamp', 'name': 'Name / Contact Person', 'phone': 'Contact Number', 'bloodGroup': 'Blood Group', 'age': 'Age',
                        'gender': 'Gender', 'Cities': 'City / Area / Location', 'Category2': 'State / Territory', 'selectDistrict': 'District', 'email': 'Email Id'}, inplace=True)
    df2.fillna(0)
    df = df2.to_excel('./plasmadonoroutput.xlsx', sheet_name='Sheet1')
    exp = pxl.load_workbook('plasmadonoroutput.xlsx')
    sheet = exp['Sheet1']
    st.write("## **Showing Plasma Donors**")
    page_name = ['Choose Option', 'View Full Data',
                 'View Filtered Data']
    page = st.radio('Choose your preferred type', page_name)
    if page == "Choose Option":
        st.info("👆 You have to choose any of the above two Options.")
    elif page == "View Full Data":
        df2 = df2.iloc[::-1]
        st.write(df2.iloc[:, 1:10])
    elif page == "View Filtered Data":
        st.write(
            "Search the Plasma Donors by District or Blood Group or Combined")
        option_t = st.checkbox('Blood Group')
        option_u = st.checkbox('District')
        if(option_t == True and option_u == True):
            district = st.selectbox(
                "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                    "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                    ))
            if district == "Not Listed":
                district = st.text_input('Enter Your District Name')
            bg = st.selectbox("Enter Blood Group", ("Select Blood Group",
                                                    "A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"))
            list = []
            for i in range(sheet.max_row+1, 2, -1):
                if (str(sheet.cell(row=i, column=5).value).upper() == bg.upper()) and (str(sheet.cell(row=i, column=9).value).upper() == district.upper()):
                    list.append(i-2)
            st.write(df2.iloc[list, 1:11])
            list = []
        if(option_t == True and option_u == False):

            bg = st.selectbox("Enter Blood Group", ("Select Blood Group",
                                                    "A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"))
            list = []
            for i in range(sheet.max_row+1, 2, -1):
                if (str(sheet.cell(row=i, column=5).value).upper() == bg.upper()):
                    list.append(i-2)
            st.write(df2.iloc[list, 1:11])
            list = []
        if(option_t == False and option_u == True):
            district = st.selectbox(
                "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                    "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                    ))
            if district == "Not Listed":
                district = st.text_input('Enter Your District Name')
            list = []
            for i in range(sheet.max_row+1, 2, -1):
                if (str(sheet.cell(row=i, column=9).value).upper() == district.upper()):
                    list.append(i-2)
            st.write(df2.iloc[list, 1:11])
            list = []

if selected == "View Oxygen Requirements":
    gc = gspread.authorize(credentials)
    sht2 = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/11E1NjRZxsP7jS6MPZ_JnEnKFopwwTFyJOZRk-lKZEoc/edit#gid=0')
    worksheet = sht2.sheet1
    df2 = get_as_dataframe(worksheet)
    df2 = df2[df2['phone'].notna()]
    df2.rename(columns={'timestamp': 'Timestamp', 'name': 'Name / Contact Person', 'phone': 'Contact Number', 'Cities': 'City / Area / Location',
               'Category': 'State / Territory', 'selectDistrict': 'District', 'email': 'Email Id'}, inplace=True)
    df2.fillna(0)
    df = df2.to_excel('./oxygenreqoutput.xlsx', sheet_name='Sheet1')
    exp = pxl.load_workbook('oxygenreqoutput.xlsx')
    sheet = exp['Sheet1']
    st.write("## **Showing Oxygen Requirements**")
    page_name = ['Choose Option', 'View Full Data',
                 'View Filtered Data']
    page = st.radio('Choose your preferred type', page_name)
    if page == "Choose Option":
        st.info("👆 You have to choose any of the above two Options.")
    elif page == "View Full Data":
        df2 = df2.iloc[::-1]
        st.write(df2.iloc[:, 1:7])
    elif page == "View Filtered Data":
        district = st.selectbox(
            "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                ))
        if district == "Not Listed":
            district = st.text_input('Enter Your District Name')
        list = []
        for i in range(sheet.max_row+1, 2, -1):
            if (str(sheet.cell(row=i, column=7).value).upper() == district.upper()):
                list.append(i-2)
        st.write(df2.iloc[list, 1:7])
        list = []

if selected == "View Oxygen Suppliers/Availability":
    gc = gspread.authorize(credentials)
    sht2 = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/1CP-vN1dU_rfn9BvcntAT15aFKIybyXGvZuXFMsFeGBk/edit#gid=0')
    worksheet = sht2.sheet1
    df2 = get_as_dataframe(worksheet)
    df2 = df2[df2['phone'].notna()]
    df2.rename(columns={'timestamp': 'Timestamp', 'name': 'Name / Contact Person', 'phone': 'Contact Number', 'Cities': 'City / Area / Location',
               'Category2': 'State / Territory', 'selectDistrict': 'District', 'email': 'Email Id'}, inplace=True)
    df2.fillna(0)
    df = df2.to_excel('./oxygensupplyoutput.xlsx', sheet_name='Sheet1')
    exp = pxl.load_workbook('oxygensupplyoutput.xlsx')
    sheet = exp['Sheet1']
    st.write("## **Showing Oxygen Suppliers**")
    page_name = ['Choose Option', 'View Full Data',
                 'View Filtered Data']
    page = st.radio('Choose your preferred type', page_name)
    if page == "Choose Option":
        st.info("👆 You have to choose any of the above two Options.")
    elif page == "View Full Data":
        df2 = df2.iloc[::-1]
        st.write(df2.iloc[:, 1:7])
    elif page == "View Filtered Data":
        district = st.selectbox(
            "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                ))
        if district == "Not Listed":
            district = st.text_input('Enter Your District Name')
        list = []
        for i in range(sheet.max_row+1, 2, -1):
            if (str(sheet.cell(row=i, column=6).value).upper() == district.upper()):
                list.append(i-2)
        st.write(df2.iloc[list, 1:7])
        list = []
if selected == "Search for Bed Availability":
    gc = gspread.authorize(credentials)
    sht2 = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/1nuMaTQSXx6KcBisg3mgwMl5xh5gq8ClWI_6iYZdfC1Y/edit#gid=0')
    worksheet = sht2.sheet1
    df2 = get_as_dataframe(worksheet)
    df2 = df2[df2['phone'].notna()]
    df2.rename(columns={'timestamp': 'Timestamp', 'name': 'Name / Contact Person', 'phone': 'Contact Number', 'beds': 'Available Beds', 'updateon': 'Data As Updated On',
                        'hospital': 'Hospital Name', 'pincode': 'Pincode', 'Category': 'State / Territory', 'Cities': 'City / Area / Location', 'selectDistrict': 'District'}, inplace=True)
    df2.fillna(0)
    df = df2.to_excel('./bedoutput.xlsx', sheet_name='Sheet1')
    exp = pxl.load_workbook('bedoutput.xlsx')
    sheet = exp['Sheet1']
    st.write("## **Showing Bed Availability**")
    page_name = ['Choose Option', 'View Full Data',
                 'View Filtered Data']
    page = st.radio('Choose your preferred type', page_name)
    if page == "Choose Option":
        st.info("👆 You have to choose any of the above two Options.")
    elif page == "View Full Data":
        df2 = df2.iloc[::-1]
        st.write(df2.iloc[:, 1:11])
    elif page == "View Filtered Data":
        district = st.selectbox(
            "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                ))
        if district == "Not Listed":
            district = st.text_input('Enter Your District Name')
        list = []
        for i in range(sheet.max_row+1, 2, -1):
            if (str(sheet.cell(row=i, column=11).value).upper() == district.upper()):
                list.append(i-2)
        st.write(df2.iloc[list, 1:11])
        list = []

if selected == "Search for Remdesivr Distributors":
    gc = gspread.authorize(credentials)
    sht2 = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/1MnGmSTxGZ30KClwqjpUH9P2Y16AxIccKGDYxcTJRaZI/edit#gid=0')
    worksheet = sht2.sheet1
    df2 = get_as_dataframe(worksheet)
    df2 = df2[df2['phone'].notna()]
    df2.rename(columns={'timestamp': 'Timestamp', 'name': 'Name / Contact Person', 'phone': 'Contact Number', 'bloodGroup': 'Blood Group', 'age': 'Age',
                        'gender': 'Gender', 'Cities': 'City / Area / Location', 'Category': 'State / Territory', 'selectDistrict': 'District', 'email': 'Email Id'}, inplace=True)
    df2.fillna(0)
    df = df2.to_excel('./remoutput.xlsx', sheet_name='Sheet1')
    exp = pxl.load_workbook('remoutput.xlsx')
    sheet = exp['Sheet1']
    st.write("## **Showing Remdevisr Distributors**")
    page_name = ['Choose Option', 'View Full Data',
                 'View Filtered Data']
    page = st.radio('Choose your preferred type', page_name)
    if page == "Choose Option":
        st.info("👆 You have to choose any of the above two Options.")
    elif page == "View Full Data":
        df2 = df2.iloc[::-1]
        st.write(df2.iloc[:, 1:7])
    elif page == "View Filtered Data":
        district = st.selectbox(
            "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                ))
        if district == "Not Listed":
            district = st.text_input('Enter Your District Name')
        list = []
        for i in range(sheet.max_row+1, 2, -1):
            if (str(sheet.cell(row=i, column=8).value).upper() == district.upper()):
                list.append(i-2)
        st.write(df2.iloc[list, 1:7])
        list = []

if selected == "Other Resources":
    gc = gspread.authorize(credentials)
    sht2 = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/1VZG2AbdWtdh2cpNWW5rOK0h1TkVcz9P5LCGlijADnoo/edit#gid=0')
    worksheet = sht2.sheet1
    df2 = get_as_dataframe(worksheet)
    df2 = df2[df2['phone'].notna()]
    df2.rename(columns={'timestamp': 'Timestamp', 'name': 'Name / Contact Person', 'phone': 'Contact Number', 'bloodGroup': 'Blood Group', 'age': 'Age',
                        'gender': 'Gender', 'Cities': 'City / Area / Location', 'Category': 'State / Territory', 'selectDistrict': 'District', 'email': 'Email Id'}, inplace=True)
    df2.fillna(0)
    df = df2.to_excel('./otherresoutput.xlsx', sheet_name='Sheet1')
    exp = pxl.load_workbook('otherresoutput.xlsx')
    sheet = exp['Sheet1']
    st.write("## **Other Resources**")
    page_name = ['Choose Option', 'View Full Data',
                 'View Filtered Data']
    page = st.radio('Choose your preferred type', page_name)
    if page == "Choose Option":
        st.info("👆 You have to choose any of the above two Options.")
    elif page == "View Full Data":
        df2 = df2.iloc[::-1]
        st.write(df2.iloc[:, 1:8])
    elif page == "View Filtered Data":
        district = st.selectbox(
            "Select District", ("Select District from the Dropdown", "24 PARAGANAS NORTH", "24 PARAGANAS SOUTH", "Alipurduar", "BANKURA", "BIRBHUM", "COOCHBEHAR", "DARJEELING", "DINAJPUR DAKSHIN", "DINAJPUR UTTAR", "HOOGHLY", "HOWRAH", "JALPAIGURI", "Jhargram", "KALIMPONG", "KOLKATA", "MALDAH", "MEDINIPUR EAST", "MEDINIPUR WEST", "MURSHIDABAD"
                                "NADIA", "PASCHIM BARDHAMAN", "PURBA BARDHAMAN", "PURULIA", "Not Listed"
                                ))
        if district == "Not Listed":
            district = st.text_input('Enter Your District Name')
        list = []
        for i in range(sheet.max_row+1, 2, -1):
            if (str(sheet.cell(row=i, column=8).value).upper() == district.upper()):
                list.append(i-2)
        st.write(df2.iloc[list, 1:8])
        list = []
